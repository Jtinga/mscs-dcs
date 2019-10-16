import os
import openpyxl
from openpyxl import Workbook
from math import ceil
from datetime import datetime
from shutil import copy

class SchoolSortingSection:

    def __init__(self):
        pass

    # Input: Excel Directory Path (string)
    # Output: boys (tuple), girls (tuple)
    def get_directory_data(self, excel_directory_path):
        section_path_list = os.listdir(excel_directory_path)  # Sections
        print("SECTION", section_path_list)

        section_list = []
        error_list = []
        # (Full Name, Average, Average of 3, Old Section, File Path)
        boys = []
        girls = []
        for section_folder in section_path_list:
            section_folder_path = os.path.join(excel_directory_path, section_folder)
            section_list.append(section_folder)
            print(os.path.join(excel_directory_path, section_folder_path))
            gender_list = os.listdir(section_folder_path)  # Genders
            for gender_folder in gender_list:
                gender_path = os.path.join(section_folder_path, gender_folder)
                student_list = os.listdir(gender_path)  # Students
                for student in student_list:
                    # st = datetime.datetime.now()
                    full_name = os.path.splitext(student)[0]  # Filename
                    average, average_of_three = self.get_file_data(os.path.join(gender_path, student))
                    if average is not None or average_of_three is not None:
                        if gender_folder.lower() == "boys":
                            boys.append([full_name, average, average_of_three, section_folder, os.path.join(gender_path, student)])
                        else:
                            girls.append([full_name, average, average_of_three, section_folder, os.path.join(gender_path, student)])
                    else:
                        error_list.append([full_name, section_folder, os.path.join(gender_path, student)])
                    # et = datetime.datetime.now()
                    # print("TIME", et - st)

        return boys, girls, section_list

    # Input: Excel File Path (string)
    # Output: Average (float), Average of Math, English and Science (float)
    def get_file_data(self, excel_file_path):
        print(excel_file_path)
        if not os.path.isfile(excel_file_path):
            # print("Error: File Not Found")
            raise FileNotFoundError

        book = openpyxl.load_workbook(excel_file_path, data_only=True)
        sheet = None
        try:
            sheet = book['SF10']
        except Exception as e:
            print(e)
            sheet = book.active
        average = float(sheet[self.find_cell_by_text(sheet, 'General Average', 'FINAL')].value)
        math = float(sheet[self.find_cell_by_text(sheet, 'Mathematics', 'FINAL')].value)
        english = float(sheet[self.find_cell_by_text(sheet, 'English', 'FINAL')].value)
        science = float(sheet[self.find_cell_by_text(sheet, 'Science ', 'FINAL')].value)
        average_of_three = (math + english + science) / 3
        return average, average_of_three

    # Input: Spreadsheet (sheet object - openpyxl)
    # Output: Intersection of row cell and column cell (string)
    def find_cell_by_text(self, sheet, row_text, column_text):
        row_name = None
        column_name = None
        for row in sheet.iter_rows():
            for entry in row:
                try:
                    if str(row_text).lower() == str(entry.value).lower() and row_name is None:
                        row_name = entry.row
                        # print(entry.value)
                    elif str(column_text).lower() == str(entry.value).lower() and column_name is None:
                        column_name = entry.column_letter
                        # print(entry.value)
                    elif row_name is not None and column_name is not None:
                        print(row_text, "+", column_text)
                        print("Cell:", column_name + str(row_name))
                        return column_name + str(row_name)
                except Exception as e:
                    print(e)
        print("Cell Not Found!")
        return None


    # Jomar
    def sorting(self, student_list):
        sorted_list = sorted(student_list, key=lambda x: (x[1], x[2]), reverse=True)
        print('Sorted', sorted_list)
        return sorted_list


    def ratio(self, boys_count, girls_count, estimation):

        if boys_count == 0:
            boys_section = 0
            girls_section = estimation
        elif girls_count == 0:
            boys_section = estimation
            girls_section = 0
        elif boys_count == 0 and girls_count == 0:
            boys_section = 0
            girls_section = 0
        elif boys_count == girls_count:
            girls_section = round(estimation / 2)
            boys_section = estimation - girls_section
        elif boys_count > girls_count:
            avg = boys_count / girls_count
            girls_section = round(estimation / ceil(avg))
            boys_section = estimation - girls_section
        else:
            avg = girls_count / boys_count
            boys_section = round(estimation / ceil(avg))
            girls_section = estimation - boys_section

        return boys_section, girls_section


    # Steven
    def section_list_file_creator(self, boys_list, girls_list, new_folder_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Section List"

        start_row = 1
        start_column = 1

        ws.cell(row=1, column=1).value = "BOYS"
        ws.cell(row=1, column=2).value = "OLD SECTION"
        ws.cell(row=1, column=3).value = "NEW SECTION"

        for boy in boys_list:
            ws.cell(row=start_row + 1, column=start_column).value = boy[0]
            ws.cell(row=start_row + 1, column=start_column + 1).value = boy[1]
            ws.cell(row=start_row + 1, column=start_column + 2).value = boy[2]
            start_row += 1

        ws.cell(row=start_row + 2, column=1).value = "GIRLS"
        ws.cell(row=start_row + 2, column=start_column + 1).value = "OLD SECTION"
        ws.cell(row=start_row + 2, column=start_column + 2).value = "NEW SECTION"

        for girl in girls_list:
            ws.cell(row=start_row + 3, column=start_column).value = girl[0]
            ws.cell(row=start_row + 3, column=start_column + 1).value = girl[1]
            ws.cell(row=start_row + 3, column=start_column + 2).value = girl[2]
            start_row += 1

        wb.save(filename=new_folder_path+"\\"+"SectionList.xlsx")


    def execute_folder(self,old_folder_path):
        if not os.path.exists(old_folder_path):
            raise NotADirectoryError
        new_folder = old_folder_path + "_" + datetime.now().strftime("%Y%m%d%H%M%S")
        os.mkdir(new_folder)
        return new_folder
    
    
    def classify(self,section_list, boys_per_section, girls_per_section, sorted_list_boys, sorted_list_girls, new_folder_path):
        counter_boys = 0
        counter_girls = 0
        boys_list = []
        girls_list = []
    
        for section in section_list:
            # Section level
            print("Section level: ", section)
            section_folder_name = self.create_section_folder(section, new_folder_path)
    
            # Insert boys into the current section in the loop
            for boy in sorted_list_boys[counter_boys:counter_boys+boys_per_section]:
                print(boy[4])
                self.copy_student_excel_file_old_to_new(boy[4], section_folder_name + "\\" + "BOYS")
                boys_list.append((boy[0], self.get_old_section(boy[4], "BOYS"), section))
            counter_boys += boys_per_section  # move on to the next batch sequence for the next section
    
            # Insert girls into the current section in the loop
            for girl in sorted_list_girls[counter_girls:counter_girls+girls_per_section]:
                print(girl[4])
                self.copy_student_excel_file_old_to_new(girl[4], section_folder_name + "\\" + "GIRLS")
                girls_list.append((girl[0], self.get_old_section(girl[4], "GIRLS"), section))
            counter_girls += girls_per_section  # move on to the next batch sequence for the next section
    
            # Call the generate new section list method
            self.section_list_file_creator(boys_list, girls_list, section_folder_name)
    
            girls_list = []
            boys_list = []
    
    
    def get_old_section(self,old_section_file_path, gender):
        index = old_section_file_path.split("\\").index(gender) - 1
        old_section_name = old_section_file_path.split("\\").pop(index)
    
        return old_section_name
    
    
    def create_section_folder(self,section_name, new_folder_path):
    
        dir_name = new_folder_path + "\\" + section_name
    
        try:
            # Create target Directory
            os.mkdir(dir_name)
            # print("Directory ", dir_name, " Created ")
            self.create_boys_girls_folder(dir_name)
    
            return dir_name
        except FileExistsError:
            print("Directory ", dir_name, " already exists")
    
    
    def create_boys_girls_folder(self,section_folder_path):
    
        section_folder_path_boys = section_folder_path + "\\" + "BOYS"
        section_folder_path_girls = section_folder_path + "\\" + "GIRLS"
    
        # BOYS
        try:
            # Create target Directory
            os.mkdir(section_folder_path_boys)
            print("Directory ", section_folder_path_boys, " Created ")
        except FileExistsError:
            print("Directory ", section_folder_path_boys, " already exists")
    
        # GIRLS
        try:
            # Create target Directory
            os.mkdir(section_folder_path_girls)
            print("Directory ", section_folder_path_girls, " Created ")
        except FileExistsError:
            print("Directory ", section_folder_path_girls, " already exists")
    
    
    def copy_student_excel_file_old_to_new(self, old_path, new_path):
        copy(old_path, new_path)
